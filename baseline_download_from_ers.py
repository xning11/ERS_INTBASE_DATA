"""
Download ERS international baseline projection data. 
https://www.ers.usda.gov/data-products/international-baseline-data/
"""

from __future__ import print_function, division
from bs4 import BeautifulSoup, SoupStrainer
from urllib.request import urlopen
import requests
import re
import os 

filename = 'InternationalBaseline'
extension = '.xlsx'

def intbase_webscrape(year='2020'):
    """
    Download ERS international baseline projection data. 
    https://www.ers.usda.gov/data-products/international-baseline-data/
    """

    url = 'https://www.ers.usda.gov/data-products/international-baseline-data/'
    u = urlopen(url)
    try:
        response = u.read().decode('utf-8')
    finally:
        u.close()

    prefix = 'https://www.ers.usda.gov'
    links = {}
    for link in BeautifulSoup(response, 'html.parser', parse_only=SoupStrainer('a')):
        pattern = r'/webdocs/DataFiles/51280/(.*)'

        if link.has_attr('href') and re.match(pattern, link['href']):
            name = link['href'].split('/')[-1]
            name = name.split('?')[0]
            links[name] = link['href'] 

    for fname, link in links.items():
        if year in str(fname):
            r = requests.get(prefix+link)
            with open(filename+extension, 'wb') as f:
                f.write(r.content)
    
            print(f"{filename}{extension} in {year} publishment has been downloaded.")



import pyexcel
import openpyxl
import xlwings 
import pandas as pd 


columnletter = 'ABCDEFGHIJKLM'

def intbase_tohar():
    """ 
    Clean and parse data to a consistent structure. 
    1. Remove the column "unaccounted loss" in sheet Cotton, which are added to total consumption.
    2. Remove the row 4 in sheet Poultry, if it has comments in row 4. 
    3. Shift the column "crush" in sheet Soybean meal and Soybean oil, and remove the column "extract rate".
    4. Create relevant empty columns in each sheet in order to keep all commodity sheets with same column size. 
    """
 
    wb = openpyxl.load_workbook(filename+extension)
    commodities = wb.sheetnames
    commodities = commodities[1:]
    # print("List of commodities in the database:", commodities)

    wb['Barley'].insert_cols(10)
    wb['Corn'].insert_cols(10)
    wb['Sorghum'].insert_cols(10)
    wb['Wheat'].insert_cols(10)

    wb['Rice'].insert_cols(8,3)
    wb['Beef'].insert_cols(8,3)
    wb['Pork'].insert_cols(8,3)

    wb['Poultry'].insert_cols(2,2)
    wb['Poultry'].insert_cols(8,3)

    if (wb['Poultry'].cell(5,1).value is None):
        wb['Poultry'].delete_rows(5)
    elif (len(wb['Poultry'].cell(5,1).value) >= 0):
        # wb['Poultry']['A5'] = wb['Poultry'].cell(5,1).value.strip(" ")
        wb['Poultry'].delete_rows(5)
        

    for c in ['Soybean meal', 'Soybean oil']:
        ssheet = wb[c]
        ssheet.delete_cols(2,2)
        ssheet.insert_cols(2,2)
        ssheet.insert_cols(10)

        # ssmaxrow = ssheet.max_row
        # ssheet.move_range("B1:B{}".format(ssmaxrow), cols=8)

        # for row in ssheet['C1:C{}'.format(ssmaxrow)]:
        #     for cell in row:
        #         cell.value = None
    
    cotsheet = wb['Cotton']
    cotsheet.insert_cols(9,2)
    maxrow = cotsheet.max_row 

    valstart = 8 
    valend = 20
    reggap = 18 

    chk = True 
    while chk:
        if cotsheet.cell(maxrow,1).value is None:
            maxrow -= 1 
        else:
            chk = False 
    # print(f'Sheet {c} size', maxrow, maxcol)

    while valend <= maxrow:

        for row_num in range(valstart,valend+1,1):
            cotsheet['N{}'.format(row_num)] =  "=SUM(G{}:H{})".format(row_num, row_num) 

        valstart += reggap
        valend += reggap

    wb.save(filename+'-ed1'+extension)


    # ------------------------------------------------------
    # re-save file to keep value of cells instead of formula
    # ------------------------------------------------------
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(filename+'-ed1'+extension)
    excel_book.save()
    excel_book.close()
    excel_app.quit()


    # ------------------------------------------------------
    # replace cotsheet column value 'cons' = 'cons' + 'loss'
    # ------------------------------------------------------
    wbed1 = openpyxl.load_workbook(filename+'-ed1'+extension, data_only=True)

    cotsheet = wbed1['Cotton']
    maxrow = cotsheet.max_row 

    valstart = 8 
    valend = 20

    while valend+6 <= maxrow:
        
        for row_num in range(valstart,valend+1,1):
            cotsheet['G{}'.format(row_num)] =  cotsheet.cell(row_num, 14).value

        valstart += reggap
        valend += reggap

    cotsheet.delete_cols(14) 

    for row in cotsheet['H1:H{}'.format(maxrow)]:
        for cell in row:
            cell.value = None

    wbed1.save(filename+'-ed2'+extension)


    # -------------------------------------------
    # create new excel file to combine all sheets 
    # -------------------------------------------
    wbed2 = openpyxl.load_workbook(filename+'-ed2'+extension,  data_only=True)

    newfile = openpyxl.Workbook()
    # sheet = newfile.active 

    for c in commodities:
        regstart = 5

        valstart = 8 
        valend = 20

        rmstart = 6
        rmgap = 5 
        
        counter = 0
                
        # create sheet for each commodity
        sheet = newfile.create_sheet(c)
        oldsheet = wbed2[c]
        
        maxrow = oldsheet.max_row 
        maxcol = oldsheet.max_column
        
        chk = True 
        while chk:
            if oldsheet.cell(maxrow,1).value is None:
                if oldsheet.cell(valstart,maxcol).value is None:
                    maxcol -= 1 
                maxrow -= 1 
            else:
                chk = False 
        # print(f'Sheet {c} size', maxrow, maxcol)
            
        sheet['A1'] = 'year'
        sheet['B1'] = 'harvest'
        sheet['C1'] = 'yield'
        sheet['D1'] = 'prod'
        sheet['E1'] = 'imp'
        sheet['F1'] = 'exp'
        sheet['G1'] = 'cons'
        sheet['H1'] = 'food'
        sheet['I1'] = 'feed'
        sheet['J1'] = 'crush'
        sheet['K1'] = 'stock'
        sheet['L1'] = 'reg'
        sheet['M1'] = 'comm'

        while valend <= maxrow:

            regcomm = oldsheet.cell(regstart, 1).value
            reg = re.split(r'\s{2,}', regcomm)[0]
            # comm = re.split(r'\s{2,}', regcomm)[1]

            for row in range(valstart,valend+1,1):
                for col in range(1,14,1):
                    column_cell = columnletter[col-1]
                    sheet[column_cell+str(row-(rmstart + rmgap*counter))] = oldsheet.cell(row,col).value

                    if sheet[column_cell+str(row-(rmstart + rmgap*counter))].value is None:
                        sheet[column_cell+str(row-(rmstart + rmgap*counter))] = 0 
                    
                    if col == 12:
                        sheet[column_cell+str(row-(rmstart + rmgap*counter))] = reg               
                    if col == 13:
                        sheet[column_cell+str(row-(rmstart + rmgap*counter))] = c      

            regstart += reggap
            valstart += reggap
            valend += reggap
            counter += 1 


    # -----------------------------------
    # save to file
    # -----------------------------------
    # emptysheet = newfile['Sheet']
    # newfile.remove(emptysheet)

    cleanfile = filename+'-clean'+extension 
    newfile.save(cleanfile)


    # -----------------------------------
    # combine all sheets into one sheet 
    # -----------------------------------
    df = []
    for i in commodities:    
        data = pd.read_excel(cleanfile, sheet_name = i, index_col=None, header=0) 
        df.append(data)

    df = pd.concat(df)

    # print(df.head())
    # print(df.comm.unique())

    # -----------------------------------
    # clean df to be HAR importable 
    # -----------------------------------
    # format column year to be 4-digit
    df['year'] = df['year'].astype(str)
    df['year'] = df['year'].str.strip()
    df['year'] = df['year'].str[:4]

    # print("List of periods in the database:", sorted(df.year.unique()))

    cols = df.columns[1:11]
    df[cols] = df[cols].apply(pd.to_numeric, errors='coerce').fillna(0).astype('float')

    # print(df.comm.unique())

    df.comm.replace({
        'Barley': 'bar',
        'Beef': 'beef',
        'Corn': 'corn',
        'Cotton': 'cot',
        'Pork': 'pork',
        'Poultry': 'poul',
        'Rice': 'rice',
        'Sorghum': 'sorg',
        'Soybeans': 'soy',
        'Soybean meal': 'soym',
        'Soybean oil': 'soyo',
        'Wheat': 'wht',
    }, inplace=True)

    # print(df.comm.unique())

    df = df[df.reg != 'WORLD']

    df.reg.replace({
        'ARGENTINA': 'arg',
        'AUSTRALIA': 'aus',
        'BRAZIL': 'bra',
        'CANADA': 'can',
        'CHINA': 'chn',
        'EGYPT': 'egy',
        'EUROPEAN UNION': 'eu',
        'EU': 'eu',
        'INDIA': 'ind',
        'INDONESIA': 'idn',
        'JAPAN': 'jpn',
        'KOREA, SOUTH': 'kor',
        'MALAYSIA': 'mys',
        'MEXICO': 'mex',
        'NEW ZEALAND': 'nzl',
        'RUSSIA': 'rus',
        'PHILIPPINES': 'phl',
        'THAILAND': 'tha',
        'UKRAINE': 'ukr',
        'UNITED STATES': 'usa',
        'USA': 'usa',
        'VIETNAM': 'vnm',
        'CENTRAL AMERICA & CARIBBEAN': 'xcb',
        'CENTRAL AM. & CARIBBEAN': 'xcb',
        '+CENT.AM+CARIB INCL CUBA': 'xcb',
        'CUBA' : 'xcb',
        'BANGLADESH': 'xas',
        'BURMA': 'xas',
        'CAMBODIA': 'xas',
        'OTHER ASIA & OCEANIA': 'xas',
        'PAKISTAN': 'xas',
        'HONG KONG': 'xas',
        'TAIWAN': 'xas',
        'OTHER SOUTH AMERICA': 'xsm',
        'IRAN': 'xme',
        'IRAQ': 'xme',
        'OTHER FORMER SOVIET UNION, 10 COUNTRIES': 'xme',
        'OTHER FSU (10)': 'xme',
        'OTHER MIDDLE EAST': 'xme',
        'SAUDI ARABIA': 'xme',
        'TURKEY': 'xme',
        'WEST AFRICAN CMTY=ECOWAS (LESS NIGERIA)': 'xaf',
        'WEST AFRICAN CMTY=ECOWAS (Less NIGERIA)': 'xaf',
        'WEST AFRICAN CMTY=ECOWAS (EXCLUDES NIGERIA)': 'xaf',
        'WEST AFRICAN CMTY=ECOWAS (INCLUDES NIGERIA)': 'xaf',
        'WEST AFRICAN CMTY=ECOWAS': 'xaf',
        'MOROCCO': 'xaf',
        'NIGERIA': 'xaf',
        'OTHER NORTH AFRICA': 'xaf',
        'OTHER SUB-SAHARAN AFRICA': 'xaf',
        'SOUTH AFRICA, REPUBLIC OF': 'xaf',
        'SOUTH AFRICA, REPUBLIC': 'xaf',
        'OTHER EUROPE': 'xeu',
    }, inplace=True)

    # print("List of regions in the database:", df.reg.unique())


    df_upd = df.groupby(['year','reg','comm'])[['prod','imp','exp','cons','food','feed','crush','stock']].sum().reset_index()

    df_upd = pd.melt(df_upd,
                    id_vars=['year', 'reg', 'comm'],
                    value_vars=[
                        'prod', 'imp', 'exp', 'cons', 'food', 'feed', 'stock',
                        'crush'
                    ],
                    var_name='use',
                    value_name='amount')

    # df_upd.loc[(df_upd.comm == 'soym' ) & (df_upd.use=='crush'), 'amount'] = 0
    # df_upd.loc[(df_upd.comm == 'soyo' ) & (df_upd.use=='crush'), 'amount'] = 0

    df_upd.use.replace({
        'prod': '1prod',
        'imp': '2imp',
        'exp': '3exp',
        'cons': '4cons',
        'food': '5food',
        'feed': '6feed',
        'stock': '7stock',
        'crush': '8crush',
    }, inplace=True)

    df_upd = df_upd.pivot_table(index=['year', 'reg', 'use'],
                            columns='comm',
                            values='amount').fillna(0).reset_index()

    # print(df_upd.head())

    df_upd.reg.replace({
        'arg': '01arg',
        'aus': '02aus',
        'bra': '03bra',
        'can': '04can',
        'chn': '05chn',
        'egy': '06egy',
        'eu':  '07eu',
        'ind': '08ind',
        'idn': '09idn',
        'jpn': '10jpn',
        'kor': '11kor',
        'mys': '12mys',
        'mex': '13mex',
        'nzl': '14nzl',
        'rus': '15rus',
        'phl': '16phl',
        'tha': '17tha',
        'ukr': '18ukr',
        'usa': '19usa',
        'vnm': '20vnm',
        'xcb': '21xcb',
        'xas': '22xas',
        'xsm': '23xsm',
        'xme': '24xme',
        'xaf': '25xaf',
        'xeu': '26xeu',
    }, inplace=True)

    regs = df_upd.reg.unique()
    regs = sorted(regs)
    # print(regs)


    df_final = pd.melt(df_upd,
                id_vars=['year', 'reg', 'use'],
                value_vars=[
                    'bar',	'beef',	'corn',	'cot',	'pork',	'poul',	'rice',	'sorg',	'soy',	'soym',	'soyo',	'wht',
                ],
                var_name='comm',
                value_name='amount')

    df_final = df_final.pivot_table(index=['year', 'use', 'comm'],
                            columns='reg',
                            values='amount').fillna(0).reset_index()

    df_final = pd.melt(df_final, 
                id_vars=['use', 'comm', 'year'],
                value_vars=regs,
                var_name='reg',
                value_name='amount')

    df_final.to_csv(filename+'2.csv', index=False)


    print(f'{filename}2.csv is ready to be imported into ViewHAR.')



    if os.path.exists(filename+'-ed1'+extension):
      os.remove(filename+'-ed1'+extension)

    if os.path.exists(filename+'-ed2'+extension):
      os.remove(filename+'-ed2'+extension)

    if os.path.exists(filename+'-clean'+extension):
      os.remove(filename+'-clean'+extension)

    if os.path.exists(filename+'-final'+extension):
      os.remove(filename+'-final'+extension)



if __name__ == '__main__':

    intbase_webscrape('2020')  # here you can change '2020' to the year of your pick, such as '2019'
    intbase_tohar()

